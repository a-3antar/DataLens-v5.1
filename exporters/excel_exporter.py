"""
exporters/excel_exporter.py
============================
تصدير التقارير إلى Excel (.xlsx).
كل بلوك جدول يذهب في sheet منفصل.
"""

import logging
from pathlib import Path

import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment,
    Border, Side,
)
from openpyxl.utils  import get_column_letter
from openpyxl.chart  import BarChart, LineChart, PieChart, AreaChart, Reference
from openpyxl.chart.marker import Marker
from openpyxl.drawing.image import Image as XLImage

from exporters.report_manager import ReportManager
from exporters.chart_renderer import render_gauge

logger = logging.getLogger(__name__)

# ── ألوان ────────────────────────────────────────────────────
COLOR_HEADER_BG  = "1E3A5F"
COLOR_HEADER_FG  = "FFFFFF"
COLOR_ROW_ALT    = "F1F5F9"
COLOR_KPI_VALUE  = "1D4ED8"
COLOR_TITLE      = "1E3A5F"
CHART_COLORS     = ["2563EB", "10B981", "F59E0B", "A855F7", "EF4444", "0EA5E9"]


class ExcelExporter:
    """
    تصدير تقرير إلى Excel.

    الاستخدام:
        exp = ExcelExporter(report_manager)
        exp.export(report_id, Path("report.xlsx"))
    """

    def __init__(self, report_manager: ReportManager):
        self.rm = report_manager

    def export(self, report_id: str, output_path: Path) -> dict:
        """
        تصدير تقرير إلى .xlsx.
        يرجع: {"ok": True, "path": "..."} أو {"ok": False, "error": "..."}
        """
        reports = self.rm.list_reports()
        report  = next((r for r in reports if r["id"] == report_id), None)
        if not report:
            return {"ok": False, "error": "التقرير غير موجود"}

        blocks = self.rm.get_blocks(report_id)

        try:
            output_path = Path(output_path)
            output_path.parent.mkdir(parents=True, exist_ok=True)

            wb = openpyxl.Workbook()
            wb.remove(wb.active)   # حذف الـ sheet الافتراضية الفارغة

            # Sheet ملخص التقرير
            self._build_summary_sheet(wb, report["title"], blocks)

            # Sheet لكل جدول بيانات
            table_blocks = [b for b in blocks if b["block_type"] == "table"]
            for i, block in enumerate(table_blocks, 1):
                sheet_name = f"جدول {i}"
                self._build_table_sheet(wb, sheet_name, block["content"])

            # Sheet لكل رسم بياني (بيانات + رسم Excel حي)
            chart_blocks = [b for b in blocks if b["block_type"] == "chart"]
            for i, block in enumerate(chart_blocks, 1):
                sheet_name = f"رسم {i}"
                self._build_chart_sheet(wb, sheet_name, block["content"])

            # Sheet لكل Gauge (صورة، لأن Excel لا يدعم رسم gauge حي)
            gauge_blocks = [b for b in blocks if b["block_type"] == "gauge"]
            for i, block in enumerate(gauge_blocks, 1):
                sheet_name = f"مقياس {i}"
                self._build_gauge_sheet(wb, sheet_name, block["content"])

            # Sheet لـ KPIs
            kpi_blocks = [b for b in blocks if b["block_type"] == "kpi"]
            if kpi_blocks:
                self._build_kpi_sheet(wb, kpi_blocks)

            wb.save(str(output_path))
            logger.info("Excel exported: %s", output_path)
            return {"ok": True, "path": str(output_path)}

        except Exception as e:
            logger.error("Excel export error: %s", e)
            return {"ok": False, "error": str(e)}

    # ──────────────────────────────────────────────────────────
    #  بناء الـ Sheets
    # ──────────────────────────────────────────────────────────

    def _build_summary_sheet(self, wb, title: str, blocks: list) -> None:
        """Sheet ملخص يحتوي العنوان والفقرات."""
        ws = wb.create_sheet("ملخص")
        ws.sheet_view.rightToLeft = True   # RTL

        row = 1
        # عنوان التقرير
        ws.cell(row, 1, title)
        ws.cell(row, 1).font = Font(
            bold=True, size=16,
            color=COLOR_TITLE,
        )
        ws.cell(row, 1).alignment = Alignment(horizontal="center")
        ws.merge_cells(f"A{row}:H{row}")
        row += 2

        # الفقرات والـ KPIs والـ Gauges
        for block in blocks:
            btype   = block.get("block_type")
            content = block.get("content", {})

            if btype == "paragraph":
                text = content.get("text", "")
                for line in text.splitlines():
                    if line.strip():
                        ws.cell(row, 1, line.strip())
                        ws.cell(row, 1).alignment = Alignment(horizontal="right")
                        row += 1

            elif btype == "kpi":
                label  = content.get("label", "KPI")
                actual = content.get("actual_value", 0)
                target = content.get("target_value", 0)
                unit   = content.get("unit", "")
                ws.cell(row, 1, label)
                ws.cell(row, 2, f"{actual} {unit}".strip())
                ws.cell(row, 3, f"الهدف: {target} {unit}".strip())
                ws.cell(row, 2).font = Font(bold=True, color=COLOR_KPI_VALUE)
                row += 1

            elif btype == "gauge":
                label   = content.get("label", "Gauge")
                current = content.get("current_value", 0)
                mn      = content.get("min_value", 0)
                mx      = content.get("max_value", 100)
                pct     = ((current - mn) / (mx - mn) * 100) if mx != mn else 0
                ws.cell(row, 1, label)
                ws.cell(row, 2, f"{current:,.0f} ({pct:.1f}%)")
                ws.cell(row, 3, "انظر شيت المقياس المخصص")
                ws.cell(row, 3).font = Font(italic=True, color="64748B")
                row += 1

            elif btype == "chart":
                title = content.get("title", "رسم بياني")
                ws.cell(row, 1, f"📊 {title}")
                ws.cell(row, 2, "انظر الشيت المخصص لهذا الرسم")
                ws.cell(row, 1).font = Font(italic=True, color="64748B")
                row += 1

        self._auto_width(ws)

    def _build_table_sheet(self, wb, sheet_name: str, content: dict) -> None:
        """Sheet بيانات جدول."""
        data    = content.get("data", [])
        columns = content.get("columns", [])
        if not data or not columns:
            return

        ws = wb.create_sheet(sheet_name)
        ws.sheet_view.rightToLeft = True

        # Header
        header_fill = PatternFill("solid", fgColor=COLOR_HEADER_BG)
        header_font = Font(bold=True, color=COLOR_HEADER_FG, size=10)
        thin        = Side(style="thin", color="CBD5E1")
        border      = Border(left=thin, right=thin, top=thin, bottom=thin)

        for col_idx, col_name in enumerate(columns, 1):
            cell = ws.cell(1, col_idx, str(col_name))
            cell.fill      = header_fill
            cell.font      = header_font
            cell.border    = border
            cell.alignment = Alignment(horizontal="center")

        # البيانات
        alt_fill = PatternFill("solid", fgColor=COLOR_ROW_ALT)
        for row_idx, row_data in enumerate(data, 2):
            fill = alt_fill if row_idx % 2 == 0 else None
            for col_idx, col_name in enumerate(columns, 1):
                val  = row_data.get(col_name, "")
                cell = ws.cell(row_idx, col_idx, val)
                if fill:
                    cell.fill = fill
                cell.border    = border
                cell.alignment = Alignment(horizontal="center")

        ws.freeze_panes = "A2"   # تجميد الـ header
        self._auto_width(ws)

    def _build_chart_sheet(self, wb, sheet_name: str, content: dict) -> None:
        """Sheet يحتوي بيانات الرسم + رسم بياني حي (Excel Chart) يمكن تعديله."""
        data   = content.get("data", [])
        x_col  = content.get("x_col", "")
        y_cols = content.get("y_cols", [])
        ctype  = content.get("chart_type", "bar")
        title  = content.get("title", sheet_name)

        if not data or not x_col or not y_cols:
            return

        ws = wb.create_sheet(sheet_name)
        ws.sheet_view.rightToLeft = True

        # ── كتابة جدول البيانات المصدر (يُستخدم كمصدر مباشر للرسم) ──
        header_fill = PatternFill("solid", fgColor=COLOR_HEADER_BG)
        header_font = Font(bold=True, color=COLOR_HEADER_FG, size=10)
        thin        = Side(style="thin", color="CBD5E1")
        border      = Border(left=thin, right=thin, top=thin, bottom=thin)

        columns = [x_col] + list(y_cols)
        for col_idx, col_name in enumerate(columns, 1):
            cell = ws.cell(1, col_idx, str(col_name))
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = Alignment(horizontal="center")

        for row_idx, row_data in enumerate(data, 2):
            for col_idx, col_name in enumerate(columns, 1):
                cell = ws.cell(row_idx, col_idx, row_data.get(col_name, ""))
                cell.border = border
                cell.alignment = Alignment(horizontal="center")

        self._auto_width(ws)
        n_rows = len(data) + 1  # شامل الهيدر

        # ── بناء الرسم البياني الحي ──
        chart_map = {
            "bar": BarChart, "line": LineChart,
            "pie": PieChart, "area": AreaChart,
            "scatter": LineChart,  # تقريب: خط بدون تعبئة + علامات فقط
        }
        ChartClass = chart_map.get(ctype, BarChart)
        chart = ChartClass()
        chart.title = title
        chart.style = 10
        chart.height = 9
        chart.width = 18
        if hasattr(chart, "type") and ctype == "bar":
            chart.type = "col"

        cats = Reference(ws, min_col=1, min_row=2, max_row=n_rows)
        data_ref = Reference(ws, min_col=2, min_row=1, max_col=1 + len(y_cols), max_row=n_rows)
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats)

        # تلوين السلاسل ليطابق ثيم التطبيق، وتحويلها لعلامات فقط لو scatter
        for i, series in enumerate(chart.series):
            color = CHART_COLORS[i % len(CHART_COLORS)]
            if ctype == "pie":
                break  # PieChart يُلوَّن حسب النقاط لا السلاسل
            if ctype == "scatter":
                series.marker = Marker(symbol="circle", size=7)
                series.marker.graphicalProperties.solidFill = color
                series.graphicalProperties.line.noFill = True
            elif ctype in ("bar", "area"):
                series.graphicalProperties.solidFill = color
            else:  # line
                series.graphicalProperties.line.solidFill = color
                series.graphicalProperties.line.width = 22000

        anchor_col = get_column_letter(len(columns) + 2)
        ws.add_chart(chart, f"{anchor_col}2")

    def _build_gauge_sheet(self, wb, sheet_name: str, content: dict) -> None:
        """Sheet يحتوي صورة Gauge (Excel لا يملك نوع رسم gauge أصلي)."""
        current = content.get("current_value", 0)
        mn      = content.get("min_value", 0)
        mx      = content.get("max_value", 100)
        label   = content.get("label", sheet_name)

        ws = wb.create_sheet(sheet_name)
        ws.sheet_view.rightToLeft = True

        try:
            img_buf = render_gauge(current, mn, mx, label=label, width=900, height=560)
            xl_img = XLImage(img_buf)
            xl_img.width = 630
            xl_img.height = 392
            ws.add_image(xl_img, "B2")
        except Exception as e:
            logger.error("Gauge sheet image error: %s", e)
            ws.cell(1, 1, f"{label}: {current} (النطاق {mn}–{mx})")

        # قيم خام أسفل الصورة لسهولة إعادة الاستخدام في صيغ Excel
        ws.cell(22, 2, "القيمة الحالية")
        ws.cell(22, 3, current)
        ws.cell(23, 2, "الحد الأدنى")
        ws.cell(23, 3, mn)
        ws.cell(24, 2, "الحد الأقصى")
        ws.cell(24, 3, mx)

    def _build_kpi_sheet(self, wb, kpi_blocks: list) -> None:
        """Sheet مخصص لـ KPIs."""
        ws = wb.create_sheet("المؤشرات")
        ws.sheet_view.rightToLeft = True

        headers = ["المؤشر", "القيمة الفعلية", "الهدف", "الوحدة", "الفجوة"]
        header_fill = PatternFill("solid", fgColor=COLOR_HEADER_BG)
        header_font = Font(bold=True, color=COLOR_HEADER_FG)

        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(1, col_idx, h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        for row_idx, block in enumerate(kpi_blocks, 2):
            c      = block["content"]
            actual = c.get("actual_value", 0)
            target = c.get("target_value", 0)
            gap    = actual - target
            ws.cell(row_idx, 1, c.get("label", ""))
            ws.cell(row_idx, 2, actual)
            ws.cell(row_idx, 3, target)
            ws.cell(row_idx, 4, c.get("unit", ""))
            cell_gap = ws.cell(row_idx, 5, gap)
            # تلوين الفجوة: أخضر إيجابي، أحمر سلبي
            cell_gap.font = Font(
                color="166534" if gap >= 0 else "991B1B",
                bold=True,
            )

        self._auto_width(ws)

    def _auto_width(self, ws) -> None:
        """
        ضبط عرض الأعمدة تلقائياً.

        نتجنب الاعتماد على ws.columns مباشرة لأن بعض الصفوف تحتوي
        خلايا مدمجة (MergedCell عبر merge_cells في _build_summary_sheet)،
        وهذا النوع لا يملك خاصية .column قابلة للاستخدام في كل إصدارات
        openpyxl، مما قد يرمي استثناءً ويوقف تصدير الملف بالكامل. بدل
        ذلك نمر على كل الخلايا عبر iter_rows() ونتجاهل أي MergedCell
        أو خلية فارغة بأمان.
        """
        from openpyxl.cell.cell import MergedCell

        widths: dict[str, int] = {}
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell, MergedCell) or cell.value is None:
                    continue
                col_letter = get_column_letter(cell.column)
                widths[col_letter] = max(widths.get(col_letter, 0), len(str(cell.value)))

        for col_letter, max_len in widths.items():
            ws.column_dimensions[col_letter].width = min(max_len + 4, 40)
